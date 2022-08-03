from openpyxl import Workbook     # pip install openpyxl
import time                       # 時間
from openpyxl import load_workbook
import urllib.request as httplib     # 網路
import json
from matplotlib import pyplot as plt # 繪圖

import ssl
###############################

def file_write(fileName='workfile.txt',content=""):
    fr = open(fileName, 'w', encoding="utf-8")
    fr.write(content)
    fr.close()


def openpyxl_open(iFileName):
    wb = load_workbook(iFileName)  # '新竹縣美食資料.xlsx')  # 讀取檔案
    # 方法一打開第一個 工作表單
    sheet = wb.active  # 打開一個工作欄
    return sheet

# 取得某一筆的資料  （第二筆）  方法二
def openpyxl_GetRow(sheet,row1=2):
    list1=[]
    col1=1
    while col1<=sheet.max_column:
        x=sheet.cell(row=row1, column=col1).value  # 取得資料 A2
        # print(x)
        list1.append(x)
        col1=col1+1
    return list1

# 把ROW 所有資料 加到LIST
def openpyxl_AddAllRowToList(sheet,contacts):
    print("全部的筆數：", sheet.max_row)
    n = 2
    while n <= sheet.max_row:
        #list1 = mylibs.openpyxl_GetRow(sheet, row1=n)
        list1 = openpyxl_GetRow(sheet, row1=n)
        if (list1[0] != None):
            str1 = (list1[0], list1[1], list1[2], list1[3])
            contacts.append(str1)
        n = n + 1
    return contacts


"""
wb = load_workbook('新竹縣美食資料.xlsx')  # 讀取檔案
# 方法一打開第一個 工作表單
sheet = wb.active                 # 打開一個工作欄

# 方法二 依照名稱 打開工作表單   (注意：工作表單名稱，請使用英文， 中文會 出現 警告訊息)

sheetnames = wb.get_sheet_names()
print(sheetnames)
# sheet = wb.get_sheet_by_name("新竹縣美食資料")
sheet = wb.get_sheet_by_name(sheetnames[0])

#######

sheet['A1'] = 87                  # 設定資料   A1
sheet.cell(row=1, column=2).value = 'OpenPyxl Tutorial' # 設定資料 B1
wb.save("sample_file.xlsx")


"""


# 取得某一筆的資料  （第二筆）  方法二
def xlsxGetRow(sheet,row1=2):
    list1=[]
    col1=1
    while col1<sheet.max_column:
        x=sheet.cell(row=row1, column=col1).value  # 取得資料 A2
        print(x)
        list1.append(x)
        col1=col1+1
    return list1



# 取得某一欄位的所有資料

def xlsxGetCol(sheet,col1=2):
    list1=[]
    row1 = 1
    while row1 < sheet.max_row:
        x = sheet.cell(row=row1, column=col1).value  # 取得資料 A2
        print(x)
        list1.append(x)
        row1 = row1 + 1
    return list1

# 繪製 6個表格
# 使用方法:
# ListLabel=["金價歷年走勢圖（美元／英兩）","金價歷年走勢圖（美元／英兩）"]
# mylibs.matplotlib_draw6Charts(yearList,capitalList,ListLabel)
def matplotlib_draw6Charts(yearList,goldPriceList,ListLabel):
    # 圖形繪製

    plt.rcParams['font.sans-serif'] = ['Microsoft JhengHei']  # 更換中文字型
    plt.rcParams['axes.unicode_minus'] = False  # 解決座標軸負數的負號顯示問題

    listlabels1 = ListLabel[0]  # "金價歷年走勢圖（美元／英兩）"

    fig, ax = plt.subplots(figsize=(14, 8), nrows=2, ncols=3)  # 設定視窗 寬14英吋長8英吋 上下二分 左右三分

    # 1 號 線型圖
    ax[0, 0].plot(yearList, goldPriceList, label=ListLabel[1] )  #"金價歷年走勢圖（美元／英兩）"
    ax[0, 0].legend()

    # 2 號 柱狀圖
    ax[0, 1].bar(yearList, goldPriceList, label="金價歷年走勢圖（美元／英兩）",
                 alpha=0.5,
                 width=0.2
                 )
    ax[0, 1].legend()

    # 3 號 星星圖
    ax[0, 2].plot(yearList, goldPriceList, "g*", label="金價歷年走勢圖（美元／英兩）")
    ax[0, 2].legend()

    # 4 號 階梯圖

    ax[1, 0].step(yearList, goldPriceList,
                  label="金價歷年走勢圖（美元／英兩）")
    ax[1, 0].legend()

    # 5 號 點狀圖
    ax[1, 1].scatter(yearList, goldPriceList,
                  label="金價歷年走勢圖（美元／英兩）")
    ax[1, 1].legend()

    # 6 號 填充多邊形
    ax[1, 2].fill(yearList, goldPriceList,
                  label="金價歷年走勢圖（美元／英兩）")
    ax[1, 2].legend()

    plt.show()

# 取得網路的資料 轉為 JSON 檔案
# 使用方法:
# urlEconomy = "https://apiservice.mol.gov.tw/OdService/download/A17000000J-030243-YTl"
# contentsEconomy =mylibs.url_GetToJson(urlEconomy)  # 網路下在後的字串，轉成JSON/ Dict
def url_GetToJson(url):
    ssl._create_default_https_context = ssl._create_unverified_context  # 因.urlopen發生問題，將ssl憑證排除

    req = httplib.Request(url, data=None, headers={
        'User-Agent': "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/100.0.4896.88 Safari/537.36"})
    response = httplib.urlopen(req)  # 把資料放入response
    if response.code == 200:  # 當連線狀況為 200 (正常)
        contentsFinance = response.read()  # 讀取網頁內容
        contentsFinance = contentsFinance.decode("utf8")
        # print(contentsFinance)

    contentsFinance = json.loads(contentsFinance)
    return contentsFinance


def url_Get(url):
    #  SSL  處理，  https    SSSSSS 就需要加上以下2行
    ssl._create_default_https_context = ssl._create_unverified_context  # 因.urlopen發生問題，將ssl憑證排除
    req = httplib.Request(url, data=None,
                          headers={
                              'User-Agent': "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/100.0.4896.88 Safari/537.36"})
    reponse = httplib.urlopen(req)  # 開啟連線動作
    if reponse.code == 200:  # 當連線正常時
        contents = reponse.read()  # 讀取網頁內容
        contents = contents.decode("utf-8")  # 轉換編碼為 utf-8
        return contents


